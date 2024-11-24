import os
import openpyxl
import shutil
import random
import time
import string
from zhipuai import ZhipuAI
from openai import OpenAI
from http import HTTPStatus
import dashscope
from dashscope import Generation
from dashscope.api_entities.dashscope_response import Role

####### Zhipu ####### 
def ZhipuResponse(prompt,Method_addition):
    mykey = 'zhipu key'
    zhipuClient = ZhipuAI(api_key= mykey)
    prompt = Method_addition+' \n'+prompt
    response = zhipuClient.chat.completions.create(
        model = 'glm-4',
        messages = [
            {'role':'system','content':"You are a rational decision maker."},
            {'role':'user','content':prompt}
        ],
        temperature = 0
    )
    answer = response.choices[0].message.content
    print('\nZhipu: ',answer)
    return answer

####### Qwen ####### 
dashscope.api_key = 'dashscope key' 
def QwenResponse(prompt,Method_addition):
    prompt = Method_addition+' \n'+prompt
    messages = [{'role':Role.SYSTEM,'content':'Your are a helpful assistant.'},
                {'role':Role.USER,'content':prompt}]
    response = Generation.call(
        Generation.Models.qwen_turbo,
        messages = messages,
        seed = random.randint(1,10000),
        result_format = 'message',
        temperature = 0
    )
    
    if response.status_code == HTTPStatus.OK:
        answer = response.output.choices[0]['message']['content']
        print('\nQwen: ',answer)
        return answer
    else:
        print('Request id: %s, Status id: %s, error code: %s,error message: %s'%(
        response.request_id,response.status_code,
        response.code,response.message))
        exit()

####### Llama Family: Atom-7B-Chat ####### 
llamaClient = OpenAI(
    api_key = "llama key", 
    base_url = "https://api.atomecho.cn/v1",
    )
def LlamaResponse(prompt,Method_addition):
    prompt = Method_addition+' \n'+prompt
    response = llamaClient.chat.completions.create(
        model = "Atom-7B-Chat",
        messages = [
            {"role":"system","content":"You are a helpful assistant."},
            {"role":"user","content":prompt},
        ],
        temperature=0,
        seed=random.randint(1,10000),
        stream=False,
    )
    answer = response.choices[0].message.content
    print('\nAtom-7B-Chat: ',answer)
    time.sleep(3) # limit of llama family
    return answer

#######  main program: test trials and save results ####### 
os.chdir('C:/Users/sophi/Desktop/AIPsyResearch/2_CognitiveBiases/2_LLMCogBias_Test')
# read the prompt for bias test
QuesPrompt_set = openpyxl.load_workbook('BiasQuestionSet.xlsx')
bias = QuesPrompt_set['bias29RoT']
bias_num = 29 # trial number
Ques_pool = list(string.ascii_uppercase[4:14]) # ['D'] 4-14: E-N

model_pool = ['Zhipu','Qwen','Atom-7B-Chat']
Method_name = ['Origin','CoT','RoT','DR','OS']
MethodPrompt_set = openpyxl.load_workbook('PromptDesign.xlsx')
os.chdir('C:/Users/sophi/Desktop/AIPsyResearch/2_CognitiveBiases/2_LLMCogBias_Test/Code/AllModels')

for model_i in range(2,3):
    model_name = model_pool[model_i]
    print("model name:",model_name)
    for method_i in range(3,5):
        print("method name:",Method_name[method_i])
        answers = openpyxl.Workbook()
        sheet = answers.create_sheet(title = Method_name[method_i])
        for i in range(bias_num):
            print("Bias name:",bias['B'+str(i+2)].value)
            for j in range(len(Ques_pool)):
                promptBias = bias[Ques_pool[j]+str(i+2)].value
                print(bias['B'+str(i+2)].value,'-',j+1,"Bias sub Name:",promptBias,'\n')
                if method_i == 4: # one-shot
                    Method_addition = MethodPrompt_set['Sheet1'][string.ascii_uppercase[j+1]+str(method_i+2)].value
                else:
                    Method_addition = MethodPrompt_set['Sheet1']['B'+str(method_i+2)].value
                if model_i==0: # Zhipu
                    answer = ZhipuResponse(promptBias,Method_addition)
                elif model_i==1: # Qwen
                    answer = QwenResponse(promptBias,Method_addition)
                else: # Atom-7B
                    answer = LlamaResponse(promptBias,Method_addition)
                sheet.cell(row=i+2,column=j+4,value=answer)
            time.sleep(60) # Atom
            
        # save the result
        file_name = model_name + '_RoT29BiasSub_' + Method_name[method_i] + '.xlsx'
        answers.save(file_name)
        old_path = 'C:/Users/sophi/Desktop/AIPsyResearch/2_CognitiveBiases/2_LLMCogBias_Test/Code/AllModels/'+file_name
        target_path = 'C:/Users/sophi/Desktop/AIPsyResearch/2_CognitiveBiases/2_LLMCogBias_Test/Data_RoT/AllMethods/'
        shutil.move(old_path,target_path)   
