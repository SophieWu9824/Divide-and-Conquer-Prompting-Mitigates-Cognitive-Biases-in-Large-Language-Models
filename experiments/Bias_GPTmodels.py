
import os
import openpyxl
import shutil
import requests
import time
import string
import requests.adapters

###### GPT API #############
key_openai = os.getenv('OPENAI_API_KEY')
def GPTResponse(model_name,prompt,Method_addition):
    prompt = Method_addition+' \n'+prompt
    url="https://api.openai.com/v1/chat/completions" 
    OPENAI_API_KEY=key_openai
    header={"Content-Type": "application/json","Authorization": "Bearer " +OPENAI_API_KEY}
    data={
        "model": model_name,
        "messages": [
          {"role": "system", "content": "You are a helpful assistant."},
          {"role": "user","content": prompt}],
        "temperature":0, # 0.5
        "stream":False, 
        } 
    try:
        requests.adapters.DEFAULT_RETRIES = 15
        s = requests.session()
        s.keep_alive = False
        response=requests.post(url=url,headers=header,json=data).json()
        answer = response['choices'][0]['message']['content']
        print("OpenAI: ",answer)
        time.sleep(1)
        return answer
        
    except Exception:
        print('Error!')
        time.sleep(0.5)       

########## main program: test trials and save results ############
os.chdir('C:/Users/sophi/Desktop/AIPsyResearch/2_CognitiveBiases/2_LLMCogBias_Test')
# read the prompt for bias test
QuesPrompt_set = openpyxl.load_workbook('BiasQuestionSet.xlsx')
bias = QuesPrompt_set['bias29RoT']
bias_num = 29 # trial number
Ques_pool = list(string.ascii_uppercase[4:14]) # ['D'] 4-14: E-N

model_pool = ['gpt-3.5-turbo','gpt-4']
Method_name = ['Origin','CoT','RoT','DR','OS']
MethodPrompt_set = openpyxl.load_workbook('PromptDesign.xlsx')
os.chdir('C:/Users/sophi/Desktop/AIPsyResearch/2_CognitiveBiases/2_LLMCogBias_Test/Code/AllModels')

for model_i in range(1,2):
    print("model name:",model_pool[model_i])
    answers = openpyxl.Workbook()
    for method_i in [2,4]: #range(1,5):
        print("method name:",Method_name[method_i])
        model_name = model_pool[model_i]
        sheet = answers.create_sheet(title = Method_name[method_i])
        for i in range(bias_num):
            print("Bias name:",bias['B'+str(i+2)].value)
            for j in range(len(Ques_pool)):
                promptBias = bias[Ques_pool[j]+str(i+2)].value
                print("Bias sub Name:",promptBias,'\n')
                if method_i == 4: # one-shot
                    Method_addition = MethodPrompt_set['Sheet1'][string.ascii_uppercase[j+1]+str(method_i+2)].value
                else:
                    Method_addition = MethodPrompt_set['Sheet1']['B'+str(method_i+2)].value
                answer = GPTResponse(model_name,promptBias,Method_addition)
                sheet.cell(row=i+2,column=j+4,value=answer)
        # save the result
        file_name = model_name + '_RoT29BiasSub_' + Method_name[method_i] + '.xlsx'
        answers.save(file_name)
        old_path = 'C:/Users/sophi/Desktop/AIPsyResearch/2_CognitiveBiases/2_LLMCogBias_Test/Code/AllModels/'+file_name
        target_path = 'C:/Users/sophi/Desktop/AIPsyResearch/2_CognitiveBiases/2_LLMCogBias_Test/Data_RoT/AllMethods/'
        shutil.move(old_path,target_path)   

    

