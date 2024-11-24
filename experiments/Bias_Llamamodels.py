import os
import openpyxl
import shutil
import random
import string
from openai import OpenAI
import openpyxl.workbook

##### Llama API ##########
nvidia_api_key = os.getenv("NVIDIA_API_KEY")
client = OpenAI(
    api_key = nvidia_api_key, 
    base_url = "https://integrate.api.nvidia.com/v1",
    )
def LlamaResponse(model_name,prompt,Method_addition):
    prompt += ' ' + Method_addition
    response = client.chat.completions.create(
        model = model_name, 
        messages = [
            {"role":"system","content":"You are a helpful assistant."},
            {"role":"user","content":prompt},
        ],
        temperature=0.5, # 0.5
        seed = random.randint(1,10000),
        stream = False
    )
    answer = response.choices[0].message.content
    print('\n',model_name,': ',answer)
    return answer

#######  main program: test trials and save results ####### 
os.chdir('C:/Users/sophi/Desktop/AIPsyResearch/2_CognitiveBiases/2_LLMCogBias_Test')
# read the prompt for bias test
QuesPrompt_set = openpyxl.load_workbook('BiasQuestionSet.xlsx')
bias = QuesPrompt_set['bias29RoT']
bias_num = 29 # trial number
Ques_pool = list(string.ascii_uppercase[4:14]) # ['D'] 4-14: E-N

model_pool = ['meta/llama3-8b-instruct','meta/llama3-70b-instruct']
model_name_pool = ['Llama3_8B','Llama3_70B']
Method_name = ['Origin','CoT','RoT','DR','OS']
MethodPrompt_set = openpyxl.load_workbook('PromptDesign.xlsx')
os.chdir('C:/Users/sophi/Desktop/AIPsyResearch/2_CognitiveBiases/2_LLMCogBias_Test/Code/AllModels')

for model_i in range(3):
    print("model name:",model_pool[model_i])
    answers = openpyxl.Workbook()
    for method_i in range(4):
        print("method name:",Method_name[method_i])
        model_name = model_pool[model_i]
        sheet = answers.create_sheet(title = Method_name[method_i])
        for i in range(bias_num):
            print("Bias name:",bias['B'+str(i+2)].value)
            for j in range(len(Ques_pool)):
                promptBias = bias[Ques_pool[j]+str(i+2)].value
                print("Bias sub Name:",promptBias,'\n')
                if method_i == 4: # one-shot
                    Method_addition = MethodPrompt_set['Sheet1'][string.ascii_uppercase[j+1]+str(method_i+2)].value
                else:
                    Method_addition = MethodPrompt_set['Sheet1']['B'+str(method_i+2)].value
                answer = LlamaResponse(model_name,promptBias,Method_addition)
                sheet.cell(row=i+2,column=j+4,value=answer)
            
    # save the result
    file_name = model_name_pool[model_i] + '_RoT29Bias_sub' + '.xlsx'
    answers.save(file_name)
    old_path = 'C:/Users/sophi/Desktop/AIPsyResearch/2_CognitiveBiases/2_LLMCogBias_Test/Code/AllModels/'+file_name
    target_path = 'C:/Users/sophi/Desktop/AIPsyResearch/2_CognitiveBiases/2_LLMCogBias_Test/Data_RoT/AllMethods/'
    shutil.move(old_path,target_path)   
